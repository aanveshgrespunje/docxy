"""
Excel (xlsx) -> PDF feature.
Route: POST /api/xlsx-to-pdf
"""
import io
from flask import Blueprint, request, send_file, jsonify
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

xlsx_to_pdf_bp = Blueprint('xlsx_to_pdf', __name__)


@xlsx_to_pdf_bp.route('/api/xlsx-to-pdf', methods=['POST'])
def xlsx_to_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "No spreadsheet items uploaded"}), 400
    file = request.files['file']
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file.stream, data_only=True)
        pdf_buffer = io.BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=letter)

        for sheet in wb.worksheets:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, 750, f"Sheet Grid Matrix: {sheet.title}")
            y_pos = 710
            c.setFont("Courier", 9)

            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row[:6])
                    c.drawString(50, y_pos, row_str[:110])
                    y_pos -= 15
                    if y_pos < 50:
                        c.showPage()
                        y_pos = 750
            c.showPage()

        c.save()
        pdf_buffer.seek(0)
        return send_file(pdf_buffer, as_attachment=True, download_name="spreadsheet_converted.pdf", mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
